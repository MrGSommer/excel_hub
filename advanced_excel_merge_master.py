import streamlit as st
import pandas as pd
import io
import openpyxl

def clean_value(value, delete_enabled, custom_chars):
    if isinstance(value, str):
        unwanted = [" m2", " m3", " m", "Nicht klassifiziert", "---"]
        if delete_enabled and custom_chars.strip():
            custom_list = [x.strip() for x in custom_chars.split(",") if x.strip()]
            unwanted.extend(custom_list)
        for u in unwanted:
            value = value.replace(u, "")
    return value

def detect_header(sheet, max_rows_check=10):
    best_row_idx = None
    max_non_empty = 0
    best_header = None
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_rows_check, values_only=True), start=1):
        non_empty_count = sum(1 for cell in row if cell and str(cell).strip() != "")
        if non_empty_count > max_non_empty:
            max_non_empty = non_empty_count
            best_row_idx = idx
            best_header = [str(cell).strip() if cell else "" for cell in row]
    return best_row_idx, best_header

def app():
    st.header("Advanced Merger - Master Table")
    st.markdown("Fasst ausgewählte Arbeitsblätter einer Excel-Datei zu einer Mastertabelle zusammen.")
    
    supplement_name = st.text_input("File Supplement Name", value="default")
    delete_enabled = st.checkbox("Zeichen in Zellen entfernen")
    custom_chars = st.text_input("Zusätzliche zu löschende Zeichen (kommagetrennt)", value="")
    
    uploaded_file = st.file_uploader("Excel-Datei für Master Table Merge hochladen", type=["xlsx", "xls"], key="master_table")
    if not uploaded_file:
        return
    
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    except Exception as e:
        st.error(f"Fehler beim Laden der Arbeitsmappe: {e}")
        return
    
    sheets = wb.sheetnames
    selected_sheets = st.multiselect("Arbeitsblätter auswählen", sheets)
    if not selected_sheets:
        st.info("Bitte wählen Sie mindestens ein Arbeitsblatt aus.")
        return

    merged_data = []
    all_columns = set()
    progress_bar = st.progress(0)
    total = len(selected_sheets)
    
    for i, sheet_name in enumerate(selected_sheets):
        sheet = wb[sheet_name]
        header_row, headers = detect_header(sheet)
        if not header_row:
            st.error(f"Kein Header in '{sheet_name}' gefunden.")
            continue
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if all(cell is None for cell in row):
                continue
            row_dict = {col: clean_value(val, delete_enabled, custom_chars) for col, val in zip(headers, row)}
            row_dict["SheetName"] = sheet_name
            merged_data.append(row_dict)
            all_columns.update(row_dict.keys())
        progress_bar.progress((i + 1) / total)
    
    if not merged_data:
        st.error("Keine Daten zusammengeführt.")
        return
    
    sorted_columns = list(all_columns)
    if "SheetName" in sorted_columns:
        sorted_columns.remove("SheetName")
        sorted_columns = ["SheetName"] + sorted_columns
        
    df_master = pd.DataFrame(merged_data, columns=sorted_columns)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_master.to_excel(writer, index=False, sheet_name="MasterTable")
    output.seek(0)
    
    st.success("Master Table Merge abgeschlossen.")
    st.download_button(
        "Download Master Table Excel", 
        data=output, 
        file_name=f"{supplement_name}_merged_master_table.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
