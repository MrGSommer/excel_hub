import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl import Workbook
from excel_utils import clean_columns_values, rename_columns_to_standard

def clean_value(value, delete_enabled, custom_chars):
    if isinstance(value, str):
        unwanted = [" m2", " m3", " m", "Nicht klassifiziert", "---"]
        if delete_enabled and custom_chars.strip():
            custom_list = [x.strip() for x in custom_chars.split(",") if x.strip()]
            unwanted.extend(custom_list)
        for u in unwanted:
            value = value.replace(u, "")
    return value

def app(supplement_name, delete_enabled, custom_chars):
    st.header("Merge to Sheets")
    st.markdown("Fügt jede Excel-Datei als eigenes Blatt in eine neue Arbeitsmappe ein.")

    uploaded_files = st.file_uploader("Excel-Dateien hochladen", type=["xlsx", "xls"], key="sheets_files", accept_multiple_files=True)
    if not uploaded_files:
        return

    progress_bar = st.progress(0)
    merged_wb = Workbook()
    if "Sheet" in merged_wb.sheetnames:
        merged_wb.remove(merged_wb["Sheet"])
    total = len(uploaded_files)

    for idx, uploaded_file in enumerate(uploaded_files, start=1):
        try:
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            sheet = wb.active
            sheet_name = uploaded_file.name.split('.')[0][:30]
            new_sheet = merged_wb.create_sheet(title=sheet_name)
            for row in sheet.iter_rows(values_only=True):
                cleaned_row = [clean_value(cell, delete_enabled, custom_chars) for cell in row]
                new_sheet.append(cleaned_row)
            progress_bar.progress(idx / total)
        except Exception as e:
            st.error(f"Fehler bei {uploaded_file.name}: {e}")
            continue

    output = io.BytesIO()
    merged_wb.save(output)
    output.seek(0)

    st.success("Merge to Sheets abgeschlossen.")
    st.download_button(
        "Download Merged Sheets Excel",
        data=output,
        file_name=f"{supplement_name}_merged_output_sheets.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="sheets_download_button"
    )

if __name__ == "__main__":
    app(supplement_name="default", delete_enabled=False, custom_chars="")
