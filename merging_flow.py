import streamlit as st
import pandas as pd
import io
import openpyxl
from excel_utils import detect_header_row, rename_columns_to_standard


def app(supplement_name, delete_enabled, custom_chars):
    """
    Flow: Mengenspalten-Merger aus mehreren Dateien oder einer Datei mit mehreren Tabs
    Performance-Optimierung: Phasenweise Laden
    """
    st.header("Flow: Mengen-Spalten Merger & Master Table | BETA VERSION")

    # Initialisierung Session-State
    if "flow_file_sheets" not in st.session_state:
        st.session_state.flow_file_sheets = {}
    if "flow_all_columns" not in st.session_state:
        st.session_state.flow_all_columns = []
    if "flow_columns_loaded" not in st.session_state:
        st.session_state.flow_columns_loaded = False

    # Schritt 1: Modus wählen
    mode = st.radio(
        "Modus wählen", ["Mehrere Dateien", "Eine Datei mit mehreren Tabs"], index=0
    )

    # Upload und Seitenauswahl
    if mode == "Mehrere Dateien":
        files = st.file_uploader(
            "Excel-Dateien hochladen", type=["xlsx", "xls"], accept_multiple_files=True, key="flow_upload_multi"
        )
        if files:
            st.session_state.flow_file_sheets = {}
            for f in files:
                wb = pd.ExcelFile(f, engine="openpyxl")
                # nur erstes Sheet, direkte Zuordnung
                st.session_state.flow_file_sheets[f.name] = (f, wb.sheet_names[0])
    else:
        single = st.file_uploader(
            "Eine Excel-Datei hochladen", type=["xlsx", "xls"], key="flow_upload_single"
        )
        if single:
            wb = pd.ExcelFile(single, engine="openpyxl")
            sheets = wb.sheet_names
            chosen = st.multiselect(
                "Arbeitsblätter wählen", sheets, default=sheets, key="flow_sheet_select"
            )
            if chosen:
                st.session_state.flow_file_sheets = {}
                for sheet in chosen:
                    key = f"{single.name} - {sheet}"
                    st.session_state.flow_file_sheets[key] = (single, sheet)

    # Schritt 2: Schaltfläche zum Laden der Spaltennamen
    if st.session_state.flow_file_sheets and not st.session_state.flow_columns_loaded:
        if st.button("Spaltennamen laden", key="flow_load_columns"):
            cols = []
            exclude = ["Teilprojekt", "Geschoss", "Gebäude", "Baufeld", "eBKP-H", "Unter Terrain"]
            for key, (f, sheet) in st.session_state.flow_file_sheets.items():
                df_raw = pd.read_excel(f, sheet_name=sheet, header=None, engine="openpyxl")
                hr = detect_header_row(df_raw)
                df = pd.read_excel(f, sheet_name=sheet, header=hr, engine="openpyxl")
                cols.extend(df.columns.tolist())
            # Einzigartig und ausschliessen
            unique = [c for c in dict.fromkeys(cols) if c not in exclude]
            st.session_state.flow_all_columns = unique
            st.session_state.flow_columns_loaded = True

    # Schritt 3: Hierarchie festlegen
    if st.session_state.flow_columns_loaded:
        measures = ["Flaeche", "Laenge", "Dicke", "Hoehe", "Volumen"]
        st.markdown("### Hierarchie der Mengenspalten festlegen")
        hierarchies = {}
        for m in measures:
            hierarchies[m] = st.multiselect(
                f"Spalten für {m}", options=st.session_state.flow_all_columns, key=f"flow_{m}"
            )

        # Schritt 4: Merge und Master-Tabelle per Button
        if st.button("Flow Merge & Download", key="flow_run_merge"):
            merged_data = []
            for identifier, (f, sheet) in st.session_state.flow_file_sheets.items():
                wb = openpyxl.load_workbook(f, data_only=True)
                ws = wb[sheet]
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rd = dict(zip(headers, row))
                    # Werte bereinigen & mergen
                    for k, v in list(rd.items()):
                        rd[k] = _clean_value(v, delete_enabled, custom_chars)
                    for m, cols in hierarchies.items():
                        if not cols:
                            continue
                        val = next((rd.get(c) for c in cols if rd.get(c) not in (None, "", 0, 0.0)), None)
                        name = {
                            "Flaeche": "Fläche (m2)",
                            "Laenge": "Länge (m)",
                            "Dicke": "Dicke (m)",
                            "Hoehe": "Höhe (m)",
                            "Volumen": "Volumen (m3)"
                        }[m]
                        rd[name] = val
                    # Ursprungs-Spalten entfernen
                    used = [c for cols in hierarchies.values() for c in cols]
                    for u in used:
                        rd.pop(u, None)
                    merged_data.append(rd)

            # Master DataFrame und Download
            df_master = pd.DataFrame(merged_data)
            df_master = rename_columns_to_standard(df_master)
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                df_master.to_excel(writer, index=False, sheet_name="Master")
            out.seek(0)
            st.download_button(
                "Download Master Excel",
                data=out,
                file_name=f"{supplement_name}_flow_master.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


def _clean_value(value, delete_enabled, custom_chars):
    if isinstance(value, str):
        unwanted = [" m2", " m3", " m", "Nicht klassifiziert", "---"]
        if delete_enabled and custom_chars.strip():
            unwanted += [x.strip() for x in custom_chars.split(',') if x.strip()]
        for u in unwanted:
            value = value.replace(u, "")
    try:
        num = float(value)
        if num == 0.0:
            return None
        return num
    except:
        return value
