import streamlit as st
import pandas as pd
import io
from collections import Counter
from openpyxl.styles import PatternFill
from excel_utils import (
    clean_columns_values,
    rename_columns_to_standard,
    convert_quantity_columns,
    detect_header_row,
)

def app(supplement_name, delete_enabled, custom_chars):
    st.header("Merge to Table")
    st.markdown(
        """
        Fasst mehrere Excel-Dateien zu einer Tabelle zusammen.  
        Spalten werden nach Häufigkeit sortiert.  
        GUID-Duplikate werden erkannt und markiert.
        """
    )

    uploaded_files = st.file_uploader(
        "Excel-Dateien hochladen",
        type=["xlsx", "xls"],
        key="table_files",
        accept_multiple_files=True
    )
    if not uploaded_files:
        return

    progress = st.progress(0)
    merged_rows = []
    column_freq = Counter()
    total = len(uploaded_files)

    # 1) Lesen und Häufigkeit zählen
    for idx, file in enumerate(uploaded_files, start=1):
        try:
            # Rohdaten ohne Header laden
            df_raw = pd.read_excel(file, header=None)

            # Header-Zeile automatisch erkennen
            header_row = detect_header_row(df_raw)

            # Datei mit erkanntem Header einlesen
            df = pd.read_excel(file, header=header_row)

            column_freq.update(df.columns)
            merged_rows.extend(df.to_dict("records"))

            st.success(f"{file.name}: Header in Zeile {header_row+1} erkannt.")

        except Exception as e:
            st.error(f"Fehler bei {file.name}: {e}")
        progress.progress(idx / total)

    if not merged_rows:
        st.error("Keine gültigen Daten gefunden.")
        return

    # 2) DataFrame erstellen mit Spalten nach Häufigkeit
    cols_sorted = [col for col, _ in column_freq.most_common()]
    df = pd.DataFrame(merged_rows, columns=cols_sorted)

    # 3) Spalten umbenennen und grundlegend bereinigen
    df = rename_columns_to_standard(df)
    df = clean_columns_values(df, delete_enabled, custom_chars)

    # 3.1) Spalten-Reihenfolge anpassen
    master_cols = [
        "Teilprojekt", "Gebäude", "Baufeld", "Geschoss",
        "eBKP-H", "Umbaustatus", "Unter Terrain", "Beschreibung",
        "Material", "Typ", "Name", "Ergänzung", "ING"
    ]
    measure_cols = ["Dicke (m)", "Fläche (m2)", "Volumen (m3)", "Länge (m)", "Höhe (m)"]

    ordered = []
    for col in master_cols:
        if col in df.columns:
            ordered.append(col)
    for col in measure_cols:
        if col in df.columns:
            ordered.append(col)
    for col in df.columns:
        if col not in ordered:
            ordered.append(col)
    df = df[ordered]

    # 4) GUID-Duplikate erkennen
    dup_mask = None
    if "GUID" in df.columns:
        dup_mask = df.duplicated(subset=["GUID"], keep=False)
        dup_count = dup_mask.sum()
        if dup_count:
            st.warning(f"{dup_count} Zeilen mit doppelter GUID gefunden und markiert")
            def highlight_dup(val):
                return 'background-color: yellow' if val else ''
            styled = df.style.apply(
                lambda row: [highlight_dup(flag) for flag in dup_mask],
                axis=1
            )
        else:
            st.success("Keine GUID-Duplikate gefunden.")
            st.dataframe(df.head(15))
    else:
        st.info("Spalte 'GUID' nicht gefunden.")
        st.dataframe(df.head(15))

    # 5) Download mit Markierung im Excel
    df_export = convert_quantity_columns(df.copy())
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name=supplement_name or "Merged")
        wb = writer.book
        ws = wb[supplement_name or "Merged"]

        if dup_mask is not None and dup_mask.any():
            fill = PatternFill(fill_type="solid", fgColor="FFFF00")
            for excel_row, is_dup in enumerate(dup_mask, start=2):
                if is_dup:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=excel_row, column=col_idx).fill = fill

    out.seek(0)
    st.download_button(
        "Download Merged Table Excel",
        data=out,
        file_name=f"{supplement_name or 'merged'}_merged_table.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="table_download_button"
    )
