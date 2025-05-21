import streamlit as st
import pandas as pd
import numpy as np
from excel_utils import (
    detect_header_row,
    prepend_values_cleaning
)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io


def app(supplement_name: str, delete_enabled: bool, custom_chars: str):
    """
    Streamlit-App zum Vergleichen zweier Excel-Dateien anhand der GUID.
    Gibt die neue Datei farblich markiert zurück: geänderte Zeilen grau, geänderte Zellen gelb.
    GUID dient als Primary Key.
    """
    st.title("Excel Vergleichstool 📝")
    col1, col2 = st.columns(2)
    with col1:
        old_file = st.file_uploader("Alte Version (Excel)", type=["xls", "xlsx"], key="old_comp")
    with col2:
        new_file = st.file_uploader("Neue Version (Excel)", type=["xls", "xlsx"], key="new_comp")
    if not old_file or not new_file:
        st.info("Bitte beide Dateien hochladen, um den Vergleich zu starten.")
        return

    xls_old = pd.ExcelFile(old_file, engine="openpyxl")
    xls_new = pd.ExcelFile(new_file, engine="openpyxl")
    common = list(set(xls_old.sheet_names) & set(xls_new.sheet_names))
    if not common:
        st.error("Keine gemeinsamen Arbeitsblaetter gefunden.")
        return
    sheet = st.selectbox("Arbeitsblatt waehlen", common)

    @st.cache_data
    def load_and_clean(file, name):
        raw = pd.read_excel(file, sheet_name=name, header=None, engine="openpyxl")
        hdr = detect_header_row(raw)
        df = pd.read_excel(file, sheet_name=name, header=hdr, engine="openpyxl")
        return prepend_values_cleaning(df, delete_enabled, custom_chars)

    df_old = load_and_clean(old_file, sheet)
    df_new = load_and_clean(new_file, sheet)
    if "GUID" not in df_old.columns or "GUID" not in df_new.columns:
        st.error("Spalte 'GUID' nicht in beiden Tabellen gefunden.")
        return

    master_cols = [
        "Teilprojekt", "Gebäude", "Baufeld", "Geschoss",
        "eBKP-H", "Umbaustatus", "Unter Terrain", "Beschreibung",
        "Material", "Typ", "Name", "Ergänzung"
    ]
    measure_cols = ["Dicke (m)", "Fläche (m2)", "Volumen (m3)", "Länge (m)", "Höhe (m)"]

    # Merge
    df = df_old.merge(
        df_new, on="GUID", how="outer", suffixes=("_old", "_new"), indicator=False
    )
    # Nur vorhandene Spalten
    compare = [col for col in master_cols + measure_cols
               if f"{col}_old" in df.columns and f"{col}_new" in df.columns]
    # Diff-Flags
    diffs = np.column_stack([df[f"{c}_old"] != df[f"{c}_new"] for c in compare])
    row_changed = diffs.any(axis=1)

    # Bereite Download mit openpyxl vor
    buffer = io.BytesIO()
    # Schreibe df_new (Original neue Datei) in Excel
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # schreibe alle Spalten der neuen Version
        df_new.to_excel(writer, sheet_name=sheet, index=False)
    buffer.seek(0)

    # Lese Workbook für Formatierung
    wb = load_workbook(buffer)
    ws = wb[sheet]

    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Mapiere Spaltennamen auf Spaltenindex
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {col: i+1 for i, col in enumerate(header)}

    # Färbe Zeilen und Zellen
    for idx, changed in enumerate(row_changed, start=2):  # Data ab Zeile 2
        if not changed:
            continue
        # ganze Zeile grau
        for cell in ws[idx]:
            cell.fill = grey_fill
        # geänderte Zellen gelb
        for j, col in enumerate(compare):
            c = compare[j]
            col_name = f"{c}_new"
            if df[f"{c}_old"].iloc[idx-2] != df[f"{c}_new"].iloc[idx-2]:
                # finde Spalte im neuen Sheet (ohne _new)
                if c in col_idx:
                    cell = ws.cell(row=idx, column=col_idx[c])
                    cell.fill = yellow_fill

    # Speichere formatiertes Workbook zurück
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"vergleich_{supplement_name or sheet}.xlsx"
    st.download_button(
        "Formatiertes Excel herunterladen",
        data=out,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.success("Download bereitgestellt.")
